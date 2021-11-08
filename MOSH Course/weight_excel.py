import openpyxl as xl
from openpyxl.chart import BarChart, Reference,LineChart
from openpyxl import Workbook
from datetime import date
from openpyxl.chart.axis import DateAxis

# cll = sheet['a1']
# cllone=sheet.cll(1,1)
workbook=xl.load_workbook("weight3.xlsx")
sheet = workbook['Sheet1']
ws= workbook.active

#
kg_cell = sheet.cell(sheet.max_row,3)
date_cell = sheet.cell(sheet.max_row + 1, 1)
weight_cell = sheet.cell(sheet.max_row, 2)
today_date=input("eneter today's date:")
today_weight= float(input("eneter today's weight"))
date_cell.value=today_date
weight_cell.value=today_weight
kg_cell.value="kg"
weights= Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=2,
                   max_col=2)

chart = LineChart()
chart.title= "Weight loss"
chart.style= 2
chart.y_axis.title="Weight"
chart.y_axis.crossAx= 500
chart.x_axis= DateAxis(crossAx=100)
chart.x_axis.number_format = 'd=mmm'
chart.x_axis.majorTimeUnit= "days"
chart.x_axis.title = "Date"
# chart.add_data(weights)
chart.add_data(weights)
dates= Reference(sheet,min_col=1,min_row=2,max_row=sheet.max_row)
chart.set_categories(dates)



sheet.add_chart(chart,'d2')



workbook.save("weight3.xlsx")
